"""Phase 11A — API integration tests for Projects, Tags, multi-BKP allocations.

Covers:

* Project CRUD (create, list with/without archived, get, update, archive,
  delete cascades cost_item.project_id to NULL).
* Tag CRUD + uniqueness conflict.
* Tag assignment: project + cost_item; cross-object rejection (422).
* Cost item with multi-BKP allocations:
    - create with allocations (no bkp_code) succeeds
    - create with both bkp_code AND allocations fails 422
    - update single→multi clears bkp_code
    - update multi→single drops allocations
* Filter ``?project_id`` and ``?tag_id=...&tag_id=...``.
* XLSX export apportions multi-BKP item amounts across BKP rows.
"""

from __future__ import annotations

import io
import uuid
from decimal import Decimal

import pytest
import pytest_asyncio
from app.core.security import hash_password
from app.models.cost import BkpCode
from app.models.object import (
    Object,
    ObjectMembership,
    ObjectRole,
    ObjectType,
    Unit,
)
from app.models.user import User
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

pytestmark = pytest.mark.integration

PW = "TestPasswort-9!ABC"


# ---- Helpers ---------------------------------------------------------------


async def _mk_user(session: AsyncSession, email: str) -> User:
    u = User(
        id=uuid.uuid4(),
        email=email,
        display_name=email.split("@")[0],
        password_hash=hash_password(PW),
        is_active=True,
    )
    session.add(u)
    await session.commit()
    return u


async def _login(client: AsyncClient, email: str) -> str:
    r = await client.post("/api/v1/auth/login", json={"email": email, "password": PW})
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


def _auth(token: str, client: AsyncClient) -> dict[str, str]:
    headers = {"Authorization": f"Bearer {token}"}
    csrf = client.cookies.get("reno_csrf")
    if csrf:
        headers["X-CSRF-Token"] = csrf
    return headers


def _cookies(client: AsyncClient) -> dict[str, str]:
    csrf = client.cookies.get("reno_csrf")
    return {"reno_csrf": csrf} if csrf else {}


async def _seed_bkp(session: AsyncSession) -> None:
    session.add_all(
        [
            BkpCode(code="D", parent_code=None, level=1, label_de="Technik", is_seed=True),
            BkpCode(code="D01", parent_code="D", level=2, label_de="Heizung", is_seed=True),
            BkpCode(code="E", parent_code=None, level=1, label_de="Inneres", is_seed=True),
            BkpCode(code="E01", parent_code="E", level=2, label_de="Bad", is_seed=True),
        ]
    )
    await session.commit()


async def _mk_object(
    session: AsyncSession, owner: User, name: str = "Haus"
) -> tuple[Object, list[Unit]]:
    obj = Object(
        id=uuid.uuid4(),
        name=name,
        type=ObjectType.MFH,
        planning_horizon_years=30,
    )
    session.add(obj)
    await session.flush()
    units = [
        Unit(object_id=obj.id, label="EG", wertquote_permille=600),
        Unit(object_id=obj.id, label="OG", wertquote_permille=400),
    ]
    for u in units:
        session.add(u)
    session.add(ObjectMembership(user_id=owner.id, object_id=obj.id, role=ObjectRole.OWNER))
    await session.commit()
    return obj, units


# ---- Fixtures ---------------------------------------------------------------


@pytest_asyncio.fixture()
async def owner(db_session: AsyncSession) -> User:
    return await _mk_user(db_session, "owner-pt@example.ch")


@pytest_asyncio.fixture()
async def seed_bkp(db_session: AsyncSession) -> None:
    await _seed_bkp(db_session)


@pytest_asyncio.fixture()
async def obj_with_units(db_session: AsyncSession, owner: User) -> tuple[Object, list[Unit]]:
    return await _mk_object(db_session, owner)


# ---- Project CRUD ----------------------------------------------------------


class TestProjectCrud:
    async def test_full_lifecycle(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
        seed_bkp: None,
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        # Create
        r = await integration_client.post(
            f"/api/v1/objects/{obj.id}/projects",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={"name": "Badsanierung", "status": "planned", "planned_year": 2027},
        )
        assert r.status_code == 201, r.text
        project_id = r.json()["id"]
        assert r.json()["name"] == "Badsanierung"

        # List (no archived)
        r = await integration_client.get(
            f"/api/v1/objects/{obj.id}/projects",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        assert any(p["id"] == project_id for p in r.json())

        # Get one
        r = await integration_client.get(
            f"/api/v1/projects/{project_id}",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        assert r.json()["status"] == "planned"

        # Update
        r = await integration_client.patch(
            f"/api/v1/projects/{project_id}",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={"status": "in_progress", "description": "Notiz"},
        )
        assert r.status_code == 200
        assert r.json()["status"] == "in_progress"
        assert r.json()["description"] == "Notiz"

        # Archive
        r = await integration_client.post(
            f"/api/v1/projects/{project_id}/archive",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
        )
        assert r.status_code == 200
        assert r.json()["archived_at"] is not None

        # Not in default list
        r = await integration_client.get(
            f"/api/v1/objects/{obj.id}/projects",
            headers=_auth(token, integration_client),
        )
        assert not any(p["id"] == project_id for p in r.json())

        # In archived list
        r = await integration_client.get(
            f"/api/v1/objects/{obj.id}/projects?include_archived=true",
            headers=_auth(token, integration_client),
        )
        assert any(p["id"] == project_id for p in r.json())

    async def test_delete_sets_cost_item_project_id_null(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
        seed_bkp: None,
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        # Create project
        r = await integration_client.post(
            f"/api/v1/objects/{obj.id}/projects",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={"name": "P1"},
        )
        assert r.status_code == 201
        project_id = r.json()["id"]

        # Create cost item linked to project
        r = await integration_client.post(
            f"/api/v1/objects/{obj.id}/cost-items",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={
                "bkp_code": "D01",
                "title": "Item-with-project",
                "planned_amount_chf": "100.00",
                "scope": "shared",
                "project_id": project_id,
            },
        )
        assert r.status_code == 201, r.text
        item_id = r.json()["id"]
        assert r.json()["project_id"] == project_id

        # Delete project
        r = await integration_client.delete(
            f"/api/v1/projects/{project_id}",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
        )
        assert r.status_code == 204

        # Item still exists with project_id NULL
        r = await integration_client.get(
            f"/api/v1/objects/{obj.id}/cost-items/{item_id}",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        assert r.json()["project_id"] is None


# ---- Tag CRUD ---------------------------------------------------------------


class TestTagCrud:
    async def test_create_list_update_delete(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        # Create
        r = await integration_client.post(
            f"/api/v1/objects/{obj.id}/tags",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={"key": "phase", "value": "A", "color": "#aabbcc"},
        )
        assert r.status_code == 201, r.text
        tag_id = r.json()["id"]

        # List
        r = await integration_client.get(
            f"/api/v1/objects/{obj.id}/tags",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        assert len(r.json()) == 1

        # Update
        r = await integration_client.patch(
            f"/api/v1/tags/{tag_id}",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={"value": "B"},
        )
        assert r.status_code == 200
        assert r.json()["value"] == "B"

        # Delete
        r = await integration_client.delete(
            f"/api/v1/tags/{tag_id}",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
        )
        assert r.status_code == 204

    async def test_uniqueness_violation(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        for _ in range(1):
            r = await integration_client.post(
                f"/api/v1/objects/{obj.id}/tags",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={"key": "phase", "value": "A"},
            )
            assert r.status_code == 201, r.text

        r2 = await integration_client.post(
            f"/api/v1/objects/{obj.id}/tags",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={"key": "phase", "value": "A"},
        )
        assert r2.status_code == 409


# ---- Tag assignment --------------------------------------------------------


class TestTagAssignment:
    async def test_assign_to_project_and_cost_item(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
        seed_bkp: None,
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        # Tag
        tag = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/tags",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={"key": "quality", "value": "premium"},
            )
        ).json()

        # Project
        project = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/projects",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={"name": "P"},
            )
        ).json()

        # Cost item
        item = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/cost-items",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={
                    "bkp_code": "D01",
                    "title": "Item",
                    "planned_amount_chf": "100.00",
                    "scope": "shared",
                },
            )
        ).json()

        for target_type, target_id in [
            ("project", project["id"]),
            ("cost_item", item["id"]),
        ]:
            r = await integration_client.post(
                f"/api/v1/tags/{tag['id']}/assignments",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={"target_type": target_type, "target_id": target_id},
            )
            assert r.status_code == 201, r.text

        # List tags for project
        r = await integration_client.get(
            f"/api/v1/project/{project['id']}/tags",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        assert len(r.json()) == 1

        # Unassign
        r = await integration_client.delete(
            f"/api/v1/tags/{tag['id']}/assignments/project/{project['id']}",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
        )
        assert r.status_code == 204

    async def test_cross_object_assignment_rejected(
        self,
        integration_client: AsyncClient,
        db_session: AsyncSession,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
    ) -> None:
        obj1, _ = obj_with_units
        # Second object owned by the same user
        obj2, _ = await _mk_object(db_session, owner, name="Haus 2")
        token = await _login(integration_client, owner.email)

        # Tag in obj1
        tag = (
            await integration_client.post(
                f"/api/v1/objects/{obj1.id}/tags",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={"key": "k", "value": "v"},
            )
        ).json()
        # Project in obj2
        project = (
            await integration_client.post(
                f"/api/v1/objects/{obj2.id}/projects",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={"name": "P2"},
            )
        ).json()

        r = await integration_client.post(
            f"/api/v1/tags/{tag['id']}/assignments",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={"target_type": "project", "target_id": project["id"]},
        )
        assert r.status_code == 422


# ---- Multi-BKP cost items --------------------------------------------------


class TestMultiBkpCostItems:
    async def test_create_with_bkp_allocations(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
        seed_bkp: None,
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        r = await integration_client.post(
            f"/api/v1/objects/{obj.id}/cost-items",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={
                "title": "Refit",
                "planned_amount_chf": "1000.00",
                "scope": "shared",
                "bkp_allocations": [
                    {"bkp_code": "D", "share_permille": 600},
                    {"bkp_code": "E", "share_permille": 400},
                ],
            },
        )
        assert r.status_code == 201, r.text
        assert r.json()["bkp_code"] is None
        assert len(r.json()["bkp_allocations"]) == 2

    async def test_create_with_both_bkp_and_allocations_rejected(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
        seed_bkp: None,
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        r = await integration_client.post(
            f"/api/v1/objects/{obj.id}/cost-items",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={
                "bkp_code": "D01",
                "title": "Bad",
                "planned_amount_chf": "100.00",
                "scope": "shared",
                "bkp_allocations": [
                    {"bkp_code": "D", "share_permille": 500},
                    {"bkp_code": "E", "share_permille": 500},
                ],
            },
        )
        assert r.status_code == 422

    async def test_update_single_to_multi_clears_bkp_code(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
        seed_bkp: None,
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        # Create single-BKP item
        r = await integration_client.post(
            f"/api/v1/objects/{obj.id}/cost-items",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={
                "bkp_code": "D01",
                "title": "Item",
                "planned_amount_chf": "100.00",
                "scope": "shared",
            },
        )
        assert r.status_code == 201
        item_id = r.json()["id"]

        # Patch: send only bkp_allocations → bkp_code should auto-clear
        r = await integration_client.patch(
            f"/api/v1/objects/{obj.id}/cost-items/{item_id}",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={
                "bkp_allocations": [
                    {"bkp_code": "D", "share_permille": 500},
                    {"bkp_code": "E", "share_permille": 500},
                ],
            },
        )
        assert r.status_code == 200, r.text
        assert r.json()["bkp_code"] is None
        assert len(r.json()["bkp_allocations"]) == 2

    async def test_update_multi_to_single_drops_allocations(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
        seed_bkp: None,
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        # Create multi
        r = await integration_client.post(
            f"/api/v1/objects/{obj.id}/cost-items",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={
                "title": "Item",
                "planned_amount_chf": "100.00",
                "scope": "shared",
                "bkp_allocations": [
                    {"bkp_code": "D", "share_permille": 500},
                    {"bkp_code": "E", "share_permille": 500},
                ],
            },
        )
        assert r.status_code == 201
        item_id = r.json()["id"]

        # Switch to single: empty list + set bkp_code
        r = await integration_client.patch(
            f"/api/v1/objects/{obj.id}/cost-items/{item_id}",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={"bkp_code": "D01", "bkp_allocations": []},
        )
        assert r.status_code == 200, r.text
        assert r.json()["bkp_code"] == "D01"
        assert r.json()["bkp_allocations"] == []


# ---- Filters ---------------------------------------------------------------


class TestCostItemFilters:
    async def test_filter_by_project_id(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
        seed_bkp: None,
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        # Project + two items, one linked, one not
        project = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/projects",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={"name": "P"},
            )
        ).json()

        linked = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/cost-items",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={
                    "bkp_code": "D01",
                    "title": "Linked",
                    "planned_amount_chf": "100.00",
                    "scope": "shared",
                    "project_id": project["id"],
                },
            )
        ).json()
        unlinked = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/cost-items",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={
                    "bkp_code": "D01",
                    "title": "Free",
                    "planned_amount_chf": "100.00",
                    "scope": "shared",
                },
            )
        ).json()

        r = await integration_client.get(
            f"/api/v1/objects/{obj.id}/cost-items?project_id={project['id']}",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        ids = {i["id"] for i in r.json()}
        assert linked["id"] in ids
        assert unlinked["id"] not in ids

    async def test_filter_by_tag_id_multi(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
        seed_bkp: None,
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        tag1 = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/tags",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={"key": "k1", "value": "v"},
            )
        ).json()
        tag2 = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/tags",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={"key": "k2", "value": "v"},
            )
        ).json()

        item_a = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/cost-items",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={
                    "bkp_code": "D01",
                    "title": "A",
                    "planned_amount_chf": "100.00",
                    "scope": "shared",
                },
            )
        ).json()
        item_b = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/cost-items",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={
                    "bkp_code": "D01",
                    "title": "B",
                    "planned_amount_chf": "100.00",
                    "scope": "shared",
                },
            )
        ).json()
        item_c = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/cost-items",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={
                    "bkp_code": "D01",
                    "title": "C",
                    "planned_amount_chf": "100.00",
                    "scope": "shared",
                },
            )
        ).json()

        # Assign tag1 → A, tag2 → B; C untagged.
        await integration_client.post(
            f"/api/v1/tags/{tag1['id']}/assignments",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={"target_type": "cost_item", "target_id": item_a["id"]},
        )
        await integration_client.post(
            f"/api/v1/tags/{tag2['id']}/assignments",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={"target_type": "cost_item", "target_id": item_b["id"]},
        )

        r = await integration_client.get(
            f"/api/v1/objects/{obj.id}/cost-items?tag_id={tag1['id']}&tag_id={tag2['id']}",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        ids = {i["id"] for i in r.json()}
        assert item_a["id"] in ids
        assert item_b["id"] in ids
        assert item_c["id"] not in ids


class TestIncludeTagIds:
    """``?include_tag_ids=true`` returns per-item ``tag_ids`` in a batched query."""

    async def test_default_omits_tag_ids(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
        seed_bkp: None,
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        await integration_client.post(
            f"/api/v1/objects/{obj.id}/cost-items",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={
                "bkp_code": "D01",
                "title": "X",
                "planned_amount_chf": "100.00",
                "scope": "shared",
            },
        )
        r = await integration_client.get(
            f"/api/v1/objects/{obj.id}/cost-items",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        for item in r.json():
            assert item["tag_ids"] is None

    async def test_include_returns_tag_ids(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
        seed_bkp: None,
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        tag1 = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/tags",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={"key": "k1", "value": "v"},
            )
        ).json()
        tag2 = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/tags",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={"key": "k2", "value": "v"},
            )
        ).json()

        tagged = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/cost-items",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={
                    "bkp_code": "D01",
                    "title": "Tagged",
                    "planned_amount_chf": "100.00",
                    "scope": "shared",
                },
            )
        ).json()
        untagged = (
            await integration_client.post(
                f"/api/v1/objects/{obj.id}/cost-items",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={
                    "bkp_code": "D01",
                    "title": "Untagged",
                    "planned_amount_chf": "100.00",
                    "scope": "shared",
                },
            )
        ).json()

        for t in (tag1, tag2):
            await integration_client.post(
                f"/api/v1/tags/{t['id']}/assignments",
                headers=_auth(token, integration_client),
                cookies=_cookies(integration_client),
                json={"target_type": "cost_item", "target_id": tagged["id"]},
            )

        r = await integration_client.get(
            f"/api/v1/objects/{obj.id}/cost-items?include_tag_ids=true",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        by_id = {i["id"]: i for i in r.json()}
        assert set(by_id[tagged["id"]]["tag_ids"]) == {tag1["id"], tag2["id"]}
        assert by_id[untagged["id"]]["tag_ids"] == []


# ---- Export integration ----------------------------------------------------


class TestExportMultiBkp:
    async def test_xlsx_apportions_multi_bkp_amounts(
        self,
        integration_client: AsyncClient,
        owner: User,
        obj_with_units: tuple[Object, list[Unit]],
        seed_bkp: None,
    ) -> None:
        obj, _ = obj_with_units
        token = await _login(integration_client, owner.email)

        # Multi-BKP item: 1000 CHF split 600/400 between D/E
        await integration_client.post(
            f"/api/v1/objects/{obj.id}/cost-items",
            headers=_auth(token, integration_client),
            cookies=_cookies(integration_client),
            json={
                "title": "Refit",
                "planned_amount_chf": "1000.00",
                "scope": "shared",
                "bkp_allocations": [
                    {"bkp_code": "D", "share_permille": 600},
                    {"bkp_code": "E", "share_permille": 400},
                ],
            },
        )

        r = await integration_client.get(
            f"/api/v1/objects/{obj.id}/export/xlsx",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["Kostenpositionen"]
        # Two data rows for the same title, one per BKP share, apportioned.
        rows: list[tuple[str, str, float | None]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is None:
                continue
            rows.append((row[0], row[2], row[6]))  # (bkp, title, planned)
        refit_rows = [r for r in rows if r[1] == "Refit"]
        assert len(refit_rows) == 2
        by_bkp = {r[0]: r[2] for r in refit_rows}
        assert by_bkp["D"] == pytest.approx(600.00)
        assert by_bkp["E"] == pytest.approx(400.00)
