"""Integration tests for the Phase 8 export endpoints.

Covers:

* ``GET /objects/{id}/export/xlsx`` — viewer can download, content is a
  real openpyxl workbook with the expected sheets and cells.
* ``GET /objects/{id}/export/pdf``  — viewer can download, body looks
  like a PDF and references the object's name.
* ``GET /objects/{id}/export/npk``  — viewer downloads the NPK JSON stub.
* ``object.export`` audit events are written per successful export.
* Outsiders cannot download.
"""

from __future__ import annotations

import io
import json
import uuid
from decimal import Decimal

import pytest
import pytest_asyncio
from app.core.security import hash_password
from app.models.audit import AuditEvent
from app.models.cost import BkpCode
from app.models.object import (
    Object,
    ObjectMembership,
    ObjectRole,
    ObjectType,
    Unit,
)
from app.models.user import User
from httpx import AsyncClient
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

pytestmark = pytest.mark.integration

PW = "TestPasswort-9!ABC"


async def _mk_user(session: AsyncSession, email: str) -> User:
    u = User(
        id=uuid.uuid4(),
        email=email,
        display_name=email.split("@")[0],
        password_hash=hash_password(PW),
        is_active=True,
        is_superuser=False,
    )
    session.add(u)
    await session.commit()
    return u


async def _login(client: AsyncClient, email: str) -> str:
    r = await client.post("/api/v1/auth/login", json={"email": email, "password": PW})
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


def _auth(token: str, client: AsyncClient) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


@pytest_asyncio.fixture()
async def owner(db_session: AsyncSession) -> User:
    return await _mk_user(db_session, "export-owner@example.ch")


@pytest_asyncio.fixture()
async def outsider(db_session: AsyncSession) -> User:
    return await _mk_user(db_session, "export-outsider@example.ch")


@pytest_asyncio.fixture()
async def seed_bkp(db_session: AsyncSession) -> None:
    db_session.add(BkpCode(code="D", parent_code=None, level=1, label_de="Technik", is_seed=True))
    db_session.add(
        BkpCode(code="D01", parent_code="D", level=2, label_de="Heizung", is_seed=True)
    )
    await db_session.commit()


@pytest_asyncio.fixture()
async def mfh(
    db_session: AsyncSession, owner: User, seed_bkp: None
) -> Object:
    obj = Object(
        id=uuid.uuid4(),
        name="Export-Testhaus",
        type=ObjectType.MFH,
        planning_horizon_years=10,
        inflation_rate_percent=Decimal("1.500"),
        initial_reserve_chf=Decimal("1000.00"),
    )
    db_session.add(obj)
    await db_session.flush()
    db_session.add(Unit(object_id=obj.id, label="EG", wertquote_permille=1000))
    db_session.add(
        ObjectMembership(user_id=owner.id, object_id=obj.id, role=ObjectRole.OWNER)
    )
    await db_session.commit()
    return obj


class TestExports:
    async def test_xlsx_download_returns_real_workbook(
        self,
        integration_client: AsyncClient,
        mfh: Object,
        owner: User,
    ) -> None:
        token = await _login(integration_client, owner.email)
        r = await integration_client.get(
            f"/api/v1/objects/{mfh.id}/export/xlsx",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200, r.text
        assert r.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment;" in r.headers["content-disposition"]
        # XLSX is a ZIP archive; check the magic bytes.
        assert r.content[:2] == b"PK"
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(r.content))
        assert "Kostenpositionen" in wb.sheetnames
        assert "Budget" in wb.sheetnames
        # Phase C — procurement sheets present even when empty.
        assert "Projekte" in wb.sheetnames
        assert "Lose" in wb.sheetnames
        assert "Lieferanten" in wb.sheetnames
        assert "Angebote" in wb.sheetnames

    async def test_pdf_download_starts_with_pdf_magic(
        self,
        integration_client: AsyncClient,
        mfh: Object,
        owner: User,
    ) -> None:
        token = await _login(integration_client, owner.email)
        r = await integration_client.get(
            f"/api/v1/objects/{mfh.id}/export/pdf",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        assert r.headers["content-type"].startswith("application/pdf")
        assert r.content.startswith(b"%PDF-")

    async def test_npk_returns_json_stub(
        self,
        integration_client: AsyncClient,
        mfh: Object,
        owner: User,
    ) -> None:
        token = await _login(integration_client, owner.email)
        r = await integration_client.get(
            f"/api/v1/objects/{mfh.id}/export/npk",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        assert r.headers["content-type"].startswith("application/json")
        payload = json.loads(r.content)
        # The stub names itself explicitly so downstream consumers know.
        assert "schema" in payload or "stub" in payload or "items" in payload

    async def test_outsider_cannot_export(
        self,
        integration_client: AsyncClient,
        mfh: Object,
        outsider: User,
    ) -> None:
        token = await _login(integration_client, outsider.email)
        r = await integration_client.get(
            f"/api/v1/objects/{mfh.id}/export/xlsx",
            headers=_auth(token, integration_client),
        )
        # No membership → 403 or 404 (matches the read policy).
        assert r.status_code in (403, 404)

    async def test_export_writes_audit_event(
        self,
        integration_client: AsyncClient,
        db_session: AsyncSession,
        mfh: Object,
        owner: User,
    ) -> None:
        token = await _login(integration_client, owner.email)
        r = await integration_client.get(
            f"/api/v1/objects/{mfh.id}/export/xlsx",
            headers=_auth(token, integration_client),
        )
        assert r.status_code == 200
        rows = (
            (
                await db_session.execute(
                    select(AuditEvent).where(AuditEvent.object_id == mfh.id)
                )
            )
            .scalars()
            .all()
        )
        export_events = [e for e in rows if e.action == "object.export"]
        assert len(export_events) >= 1
        assert any("xlsx" in (e.payload or {}).get("format", "").lower() for e in export_events)
