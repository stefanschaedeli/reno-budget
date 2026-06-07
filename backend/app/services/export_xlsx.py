"""XLSX export service (Phase 8).

Builds a two-worksheet workbook for an :class:`~app.models.object.Object`:

* **Kostenpositionen** — one row per visible :class:`~app.models.cost.CostItem`
  with the fields the user expects to copy/paste into other Swiss costing
  tools (BKP code, German label, status, priority, planned/actual amounts,
  per-unit allocations, NPK stub).
* **Budget** — per-year timeline rows (planned / planned-inflated / actual)
  plus a summary section at the bottom that mirrors the Phase 4 reserve
  plan: initial reserve, future planned total, required total /
  per-year / per-month.

Money uses the Swiss ``#,##0.00`` number format; percentages use
``0.00%``. Headers are bold; the first data row stays frozen on scroll.

Security / RBAC
---------------
The caller MUST pass a resolved :class:`~app.services.rbac.ObjectAccess`.
We reuse :mod:`app.services.budgets`' pro-rating logic so scoped EDITOR /
VIEWER members see their personal share, exactly as on the dashboard. The
``scope_pro_rated`` flag is surfaced in a header cell so users (and
auditors) can tell at a glance whether the figures represent the full
object or one member's slice.
"""

from __future__ import annotations

import io
import re
import uuid
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.cost import CostItem
from app.models.object import Object, Unit
from app.repositories.bkp import list_bkp_codes
from app.repositories.cost_item import list_cost_items as repo_list_cost_items
from app.repositories.object import get_object, list_units
from app.services.bkp_shares import iter_bkp_shares
from app.services.budgets import compute_reserve_plan, compute_timeline
from app.services.rbac import ObjectAccess

# Column titles (German) for the Kostenpositionen sheet, in display order.
_COST_HEADERS: tuple[str, ...] = (
    "BKP-Code",
    "BKP-Bezeichnung",
    "Titel",
    "Status",
    "Priorität",
    "Geplant (Jahr)",
    "Geplant CHF",
    "Effektiv CHF",
    "Ausführungsdatum",
    "Einheit / Anteil",
    "NPK-Stub",
)

_BUDGET_HEADERS: tuple[str, ...] = (
    "Jahr",
    "Geplant CHF",
    "Geplant inflationsbereinigt CHF",
    "Effektiv CHF",
)

# German display labels for enum-ish values (kept short for column widths).
_STATUS_DE: dict[str, str] = {
    "idea": "Idee",
    "planned": "Geplant",
    "in_progress": "In Arbeit",
    "completed": "Abgeschlossen",
    "cancelled": "Storniert",
}

_PRIORITY_DE: dict[str, str] = {
    "low": "Tief",
    "med": "Mittel",
    "high": "Hoch",
    "urgent": "Dringend",
}

_CHF_FORMAT = "#,##0.00"
_PERCENT_FORMAT = "0.00%"
_HEADER_FONT = Font(bold=True)
_HEADER_FILL_ALIGN = Alignment(horizontal="left", vertical="center")


def _slugify(name: str) -> str:
    """Return a filesystem-safe slug for the export filename.

    Replaces anything that's not a letter, digit, underscore or hyphen with
    a hyphen; collapses runs of hyphens; trims leading/trailing hyphens. The
    output is intentionally ASCII — the German object name "Haus Müller"
    becomes ``haus-mueller`` after umlaut folding (best-effort) so the
    download works on every browser.
    """
    if not name:
        return "objekt"
    # Common German umlaut folding so the slug is readable, not "haus-m-ller".
    folded = (
        name.replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("Ä", "Ae")
        .replace("Ö", "Oe")
        .replace("Ü", "Ue")
        .replace("ß", "ss")
    )
    lowered = folded.lower()
    slug = re.sub(r"[^a-z0-9_-]+", "-", lowered).strip("-")
    return slug or "objekt"


def build_export_filename(obj_name: str, extension: str) -> str:
    """Compose ``reno-budget_<slug>_<YYYY-MM-DD>.<ext>``.

    The date stamp is today's calendar date; the extension is given without
    a leading dot. This helper is shared by the XLSX, PDF and NPK
    exporters so all three artefacts download with consistent naming.
    """
    today = date.today().isoformat()
    return f"reno-budget_{_slugify(obj_name)}_{today}.{extension.lstrip('.')}"


def _allocations_cell(item: CostItem, units_by_id: dict[uuid.UUID, Unit]) -> str:
    """Render an item's allocations as ``"EG: 400‰ | 1.OG: 600‰"``.

    We keep this readable in a spreadsheet cell rather than spreading
    allocations across many columns — three units per cost item is typical
    in our data, so a one-line summary is enough.
    """
    parts: list[str] = []
    # Stable order: sort by unit label so the same item always renders the
    # same string (helps diff-friendliness when users version the exports).
    sorted_allocs = sorted(
        item.allocations,
        key=lambda a: units_by_id[a.unit_id].label if a.unit_id in units_by_id else "",
    )
    for a in sorted_allocs:
        unit = units_by_id.get(a.unit_id)
        label = unit.label if unit is not None else str(a.unit_id)
        parts.append(f"{label}: {a.share_permille}‰")
    return " | ".join(parts)


def _scope_factor_for_item(item: CostItem, allowed: frozenset[uuid.UUID] | None) -> Decimal:
    """Pro-rating factor for a cost item under the caller's unit scope.

    Mirrors :func:`app.services.budgets._scope_factor` so the XLSX numbers
    match the dashboard exactly: unscoped callers see the raw amounts;
    scoped callers see their permille share of the item.
    """
    if allowed is None:
        return Decimal("1")
    share = Decimal("0")
    for a in item.allocations:
        if a.unit_id in allowed:
            share += Decimal(a.share_permille)
    return share / Decimal("1000")


def _format_header(ws: Worksheet, headers: tuple[str, ...]) -> None:
    """Write ``headers`` as a bold, frozen header row in ``ws``."""
    for idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_FILL_ALIGN
    # Freeze row 1 (header). Subsequent rows scroll under it.
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, headers: tuple[str, ...]) -> None:
    """Apply pragmatic column widths.

    We don't compute the optimal width per cell (openpyxl doesn't ship a
    perfect autosize); a fixed minimum sized to the longest header keeps
    the file readable without being clever.
    """
    for idx, label in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(label) + 2)


async def _write_cost_items_sheet(
    ws: Worksheet,
    session: AsyncSession,
    object_id: uuid.UUID,
    *,
    access: ObjectAccess,
    units: list[Unit],
) -> None:
    """Populate the Kostenpositionen sheet with one row per visible item.

    Scoped callers see items whose allocations intersect their unit scope;
    the planned / actual amounts are pro-rated by the in-scope permille
    share so the totals in this sheet match the Budget sheet.
    """
    _format_header(ws, _COST_HEADERS)
    _autosize(ws, _COST_HEADERS)

    units_by_id: dict[uuid.UUID, Unit] = {u.id: u for u in units}

    # eBKP-H labels: build a lookup so we don't query per-row.
    bkp_rows = await list_bkp_codes(session)
    bkp_label: dict[str, str] = {b.code: b.label_de for b in bkp_rows}

    items = list(await repo_list_cost_items(session, object_id))
    allowed = access.allowed_unit_ids

    row = 2
    for item in items:
        factor = _scope_factor_for_item(item, allowed)
        if factor == 0:
            # Scoped caller cannot see this item — skip silently. Note: this
            # is the same rule the dashboard uses, so the export stays
            # consistent with what the user sees on screen.
            continue

        # One row per (item, bkp_share). Single-BKP items emit one row at
        # 1000‰; multi-BKP items emit one row per allocation; uncategorised
        # items emit one row with BKP "—" so they remain visible.
        for code, share_permille in iter_bkp_shares(item):
            share_frac = Decimal(share_permille) / Decimal(1000)
            planned = (
                item.planned_amount_chf * factor * share_frac
                if item.planned_amount_chf is not None
                else None
            )
            actual = (
                item.actual_amount_chf * factor * share_frac
                if item.actual_amount_chf is not None
                else None
            )

            ws.cell(row=row, column=1, value=code if code is not None else "—")
            ws.cell(row=row, column=2, value=bkp_label.get(code, "") if code else "")
            ws.cell(row=row, column=3, value=item.title)
            ws.cell(
                row=row,
                column=4,
                value=_STATUS_DE.get(str(item.status), str(item.status)),
            )
            ws.cell(
                row=row,
                column=5,
                value=_PRIORITY_DE.get(str(item.priority), str(item.priority)),
            )
            ws.cell(row=row, column=6, value=item.planned_year)
            planned_value = float(planned) if planned is not None else None
            planned_cell = ws.cell(row=row, column=7, value=planned_value)
            planned_cell.number_format = _CHF_FORMAT
            actual_value = float(actual) if actual is not None else None
            actual_cell = ws.cell(row=row, column=8, value=actual_value)
            actual_cell.number_format = _CHF_FORMAT
            ws.cell(row=row, column=9, value=item.actual_date)
            ws.cell(row=row, column=10, value=_allocations_cell(item, units_by_id))
            # NPK stub: real codes will be validated against the SIA catalogue
            # in a future phase; today we just round-trip whatever the user
            # entered (may be empty).
            ws.cell(row=row, column=11, value=item.npk_code or "")
            row += 1


async def _write_budget_sheet(
    ws: Worksheet,
    session: AsyncSession,
    object_id: uuid.UUID,
    *,
    access: ObjectAccess,
    obj: Object,
) -> None:
    """Populate the Budget sheet: per-year rows + a reserve-plan summary."""
    _format_header(ws, _BUDGET_HEADERS)
    _autosize(ws, _BUDGET_HEADERS)

    timeline = await compute_timeline(session, object_id, access=access, inflated=True)
    reserve = await compute_reserve_plan(session, object_id, access=access)

    row = 2
    for r in timeline.rows:
        ws.cell(row=row, column=1, value=r.year)
        cell_planned = ws.cell(row=row, column=2, value=float(r.planned_chf))
        cell_planned.number_format = _CHF_FORMAT
        cell_planned_infl = ws.cell(row=row, column=3, value=float(r.planned_inflated_chf))
        cell_planned_infl.number_format = _CHF_FORMAT
        cell_actual = ws.cell(row=row, column=4, value=float(r.actual_chf))
        cell_actual.number_format = _CHF_FORMAT
        row += 1

    # ---- Reserve-Zusammenfassung ----
    # Leave one empty row, then a bold section header, then key/value pairs.
    row += 1
    header_cell = ws.cell(row=row, column=1, value="Reserve-Zusammenfassung")
    header_cell.font = _HEADER_FONT
    row += 1

    def _kv(label: str, value: Decimal | int | str, *, fmt: str | None = None) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = _HEADER_FONT
        rendered = float(value) if isinstance(value, Decimal) else value
        cell = ws.cell(row=row, column=2, value=rendered)
        if fmt:
            cell.number_format = fmt
        row += 1

    _kv("Beitragsmodus", str(obj.contribution_mode))
    _kv("Inflationsrate p.a.", obj.inflation_rate_percent / Decimal("100"), fmt=_PERCENT_FORMAT)
    _kv("Anfangsreserve CHF", reserve.initial_reserve_chf, fmt=_CHF_FORMAT)
    _kv(
        "Geplant gesamt (inflationsbereinigt) CHF",
        reserve.total_planned_inflated_chf,
        fmt=_CHF_FORMAT,
    )
    _kv("Benötigtes Total CHF", reserve.required_total_chf, fmt=_CHF_FORMAT)
    _kv("Benötigt pro Jahr CHF", reserve.required_per_year_chf, fmt=_CHF_FORMAT)
    _kv("Benötigt pro Monat CHF", reserve.required_per_month_chf, fmt=_CHF_FORMAT)
    _kv("Pro-rated (Anteilig)", "ja" if reserve.scope_pro_rated else "nein")


async def build_xlsx(
    session: AsyncSession,
    object_id: uuid.UUID,
    *,
    access: ObjectAccess,
) -> tuple[bytes, str]:
    """Build the export workbook and return ``(bytes, filename)``.

    The caller (the export router) streams the bytes back to the client
    with an appropriate ``Content-Disposition``. Returning the filename
    here keeps the slugging/datestamp policy in one place.
    """
    obj = await get_object(session, object_id)
    assert obj is not None  # router resolves 404 via require_object_access
    units = list(await list_units(session, object_id))

    wb = Workbook()
    # Workbook ships with a default sheet — rename and reuse it as our first
    # worksheet so the file doesn't contain a stray empty "Sheet".
    ws_items = wb.active
    assert ws_items is not None
    ws_items.title = "Kostenpositionen"
    ws_budget = wb.create_sheet("Budget")

    await _write_cost_items_sheet(ws_items, session, object_id, access=access, units=units)
    await _write_budget_sheet(ws_budget, session, object_id, access=access, obj=obj)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), build_export_filename(obj.name, "xlsx")


__all__ = ["build_export_filename", "build_xlsx"]
